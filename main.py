import contextlib
import csv
import io
import itertools
import json
import os
import random
import re
import secrets
import shutil
import sqlite3
import urllib.parse

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import auth
import crypto
import review_export
import totp

DB_PATH = os.getenv("DB_PATH", "/data/survey.db")
UPLOADS_PATH = os.getenv("UPLOADS_PATH", "/data/uploads")

# Bootstrap admin — created once on first run, then owns any pre-existing surveys.
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "admin@survey.local").strip().lower()
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin")

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg", ".pdf"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# --- panel recruitment ---
# Panel providers (Bilendi, Dynata, Cint, Toluna...) hand each respondent a
# single-use token in the entry URL and expect us to bounce them back to one of
# three return URLs with that token attached. Without the bounce the provider
# cannot credit the respondent, so a panel field is unusable without this.
PANEL_TOKEN_RE = re.compile(r"^[A-Za-z0-9_.\-]{1,128}$")
PANEL_PARAM_RE = re.compile(r"^[A-Za-z0-9_\-]{1,32}$")
PANEL_OUTCOMES = ("complete", "screenout", "quotafull")

# An assignment issued but not yet submitted still counts toward balancing for
# this long, which keeps concurrent starts from piling onto the same arm without
# letting abandoned sessions skew the arms permanently.
PENDING_ASSIGNMENT_MINUTES = 60

# mcp_app reaches back here for get_db and the ownership check, but only at call
# time, so importing it before `app` exists resolves the cycle cleanly.
from mcp_app import mcp  # noqa: E402


@contextlib.asynccontextmanager
async def lifespan(app: FastAPI):
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    os.makedirs(UPLOADS_PATH, exist_ok=True)
    init_db()
    app.mount("/uploads", StaticFiles(directory=UPLOADS_PATH), name="uploads")
    # The MCP session manager has to be running for the mounted transport to
    # answer at all; without this every call to /mcp fails with a 500 that says
    # nothing about why.
    async with mcp.session_manager.run():
        yield


app = FastAPI(lifespan=lifespan)
templates = Jinja2Templates(directory="templates")


# The MCP transport checks Host headers against DNS rebinding, so the public
# domain has to be allowed or every proxied request is refused.
def _allowed_hosts() -> list:
    from urllib.parse import urlparse
    hosts = ["localhost:8000", "127.0.0.1:8000", "localhost", "127.0.0.1"]
    public = urlparse(os.environ.get("PUBLIC_URL", "")).netloc
    if public:
        hosts.append(public)
    return hosts


from mcp.server.transport_security import TransportSecuritySettings  # noqa: E402

app.mount("/mcp", mcp.streamable_http_app(
    streamable_http_path="/", json_response=True, stateless_http=True,
    transport_security=TransportSecuritySettings(
        allowed_hosts=_allowed_hosts(),
        allowed_origins=[os.environ.get("PUBLIC_URL", "http://localhost:8000")])))


@app.middleware("http")
async def mcp_key_gate(request: Request, call_next):
    """
    Resolve the MCP caller, or refuse.

    Two ways in, one table. The header is the normal path; /mcp/k/{key} carries
    the same key as a path segment for clients that cannot set headers, and is
    stripped before the mounted app sees it, so the MCP layer stays unaware of
    how the caller authenticated.

    Note this sits in front of /mcp only. The researcher UI keeps its own login,
    and /s/{slug} stays open to respondents, who have no account here.
    """
    path = request.url.path
    if not path.startswith("/mcp"):
        return await call_next(request)

    if path.startswith("/mcp/k/"):
        key, _, rest = path[len("/mcp/k/"):].partition("/")
        request.scope["path"] = "/mcp/" + rest
        request.scope["raw_path"] = request.scope["path"].encode()
    else:
        key = request.headers.get("X-API-Key", "")

    db = get_db()
    try:
        user = auth.check_api_key(db, key)
    finally:
        db.close()
    auth.set_caller(user)
    if not user:
        return JSONResponse({"error": "missing or invalid API key"}, status_code=401)
    return await call_next(request)


# --- database ---

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            email                 TEXT UNIQUE NOT NULL,
            name                  TEXT NOT NULL,
            hashed_password       TEXT NOT NULL,
            totp_secret_encrypted TEXT,
            totp_enabled          INTEGER NOT NULL DEFAULT 0,
            backup_codes_json     TEXT,
            is_admin              INTEGER NOT NULL DEFAULT 0,
            is_active             INTEGER NOT NULL DEFAULT 1,
            created_at            TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS surveys (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            slug        TEXT UNIQUE NOT NULL,
            title       TEXT NOT NULL,
            schema_json TEXT NOT NULL,
            active      INTEGER NOT NULL DEFAULT 1,
            created_at  TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS responses (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            survey_id     INTEGER NOT NULL REFERENCES surveys(id) ON DELETE CASCADE,
            response_json TEXT NOT NULL,
            submitted_at  TEXT NOT NULL DEFAULT (datetime('now'))
        );
    """)
    # per-user ownership: additive migration (older DBs predate this column)
    survey_cols = {r[1] for r in db.execute("PRAGMA table_info(surveys)").fetchall()}
    if "owner_id" not in survey_cols:
        db.execute("ALTER TABLE surveys ADD COLUMN owner_id INTEGER REFERENCES users(id)")

    # The immutable subject an upstream SSO gate knows this person by, when
    # there is one. Null until map_borant.py links them, and never the email:
    # an address changes with an institution, and this is what has to survive
    # that change and keep someone attached to the surveys they own.
    user_cols = {r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()}
    if "borant_sub" not in user_cols:
        db.execute("ALTER TABLE users ADD COLUMN borant_sub TEXT")
        db.execute("CREATE UNIQUE INDEX IF NOT EXISTS ix_users_borant_sub "
                   "ON users(borant_sub)")

    # migrate single-pool schema to multi-pool if needed
    has_rand_pools = db.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='rand_pools'"
    ).fetchone()
    if not has_rand_pools:
        db.executescript("DROP TABLE IF EXISTS assignment_counts; DROP TABLE IF EXISTS randomization;")
    db.executescript("""
        CREATE TABLE IF NOT EXISTS rand_pools (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            survey_id      INTEGER NOT NULL REFERENCES surveys(id) ON DELETE CASCADE,
            pool_name      TEXT NOT NULL DEFAULT 'Pool',
            pool_order     INTEGER NOT NULL DEFAULT 0,
            pool_pages     TEXT NOT NULL DEFAULT '[]',
            show_count     INTEGER NOT NULL DEFAULT 1,
            condition_var  TEXT NULL,
            condition_map  TEXT NULL,
            page_order     TEXT NULL
        );
        -- legacy aggregate counter, superseded by `assignments`. Kept so that
        -- an existing deployment can be backfilled from it exactly once.
        CREATE TABLE IF NOT EXISTS assignment_counts (
            pool_id       INTEGER NOT NULL REFERENCES rand_pools(id) ON DELETE CASCADE,
            condition_key TEXT NOT NULL,
            count         INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (pool_id, condition_key)
        );
    """)
    # migrate existing rand_pools tables that predate condition columns
    cols = {r[1] for r in db.execute("PRAGMA table_info(rand_pools)").fetchall()}
    if "condition_var" not in cols:
        db.execute("ALTER TABLE rand_pools ADD COLUMN condition_var TEXT NULL")
    if "condition_map" not in cols:
        db.execute("ALTER TABLE rand_pools ADD COLUMN condition_map TEXT NULL")
    if "page_order" not in cols:
        db.execute("ALTER TABLE rand_pools ADD COLUMN page_order TEXT NULL")

    # panel recruitment config, per survey (additive)
    for col in ("panel_token_param", "panel_complete_url",
                "panel_screenout_url", "panel_quotafull_url"):
        if col not in survey_cols:
            db.execute(f"ALTER TABLE surveys ADD COLUMN {col} TEXT NULL")

    # panel token on each response, unique per survey so a token cannot be
    # spent twice. SQLite treats NULLs as distinct, so non-panel surveys are
    # unaffected by the index.
    resp_cols = {r[1] for r in db.execute("PRAGMA table_info(responses)").fetchall()}
    if "panel_token" not in resp_cols:
        db.execute("ALTER TABLE responses ADD COLUMN panel_token TEXT NULL")
    db.execute("""CREATE UNIQUE INDEX IF NOT EXISTS idx_responses_panel_token
                  ON responses (survey_id, panel_token)""")

    # Per-assignment ledger. Replaces the aggregate assignment_counts, which
    # incremented on page load and so drifted with every abandonment and reload.
    fresh_assignments = not db.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='assignments'"
    ).fetchone()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS assignments (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            pool_id       INTEGER NOT NULL REFERENCES rand_pools(id) ON DELETE CASCADE,
            condition_key TEXT NOT NULL,
            issued_at     TEXT NOT NULL DEFAULT (datetime('now')),
            completed     INTEGER NOT NULL DEFAULT 0
        );
        CREATE INDEX IF NOT EXISTS idx_assignments_pool
            ON assignments (pool_id, completed, issued_at);
    """)
    if fresh_assignments:
        _backfill_assignments(db)

    # MCP credentials. A key belongs to a person, never to the installation:
    # every model-facing call resolves to `user_id` and then goes through the
    # same _owned_survey() the web app uses, so a key reaches exactly what its
    # owner reaches. Without that binding the MCP surface would be a hole
    # straight through survey ownership.
    db.executescript("""
        CREATE TABLE IF NOT EXISTS api_keys (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id      INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            name         TEXT NOT NULL,
            key          TEXT NOT NULL UNIQUE,
            active       INTEGER NOT NULL DEFAULT 1,
            created_at   TEXT NOT NULL DEFAULT (datetime('now')),
            last_used_at TEXT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_api_keys_user ON api_keys (user_id);
    """)
    db.commit()

    _bootstrap_admin(db)
    db.close()


def _backfill_assignments(db):
    """Seed the per-assignment ledger from the old aggregate counters so that
    balancing continues from where it left off rather than restarting.

    The old counters cannot distinguish a completed response from an abandoned
    page load, so every historical count is carried over as completed. That
    slightly overstates history, and it is the closest reconstruction available.
    """
    if not db.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='assignment_counts'"
    ).fetchone():
        return
    for row in db.execute("SELECT pool_id, condition_key, count FROM assignment_counts"):
        if row["count"] > 0:
            db.executemany(
                "INSERT INTO assignments (pool_id, condition_key, completed) VALUES (?, ?, 1)",
                [(row["pool_id"], row["condition_key"])] * row["count"],
            )


def _bootstrap_admin(db):
    """Create the bootstrap admin on first run and hand it any orphan surveys.
    The admin still has to enrol in 2FA on first login (totp_enabled = 0)."""
    admin = db.execute("SELECT id FROM users WHERE is_admin = 1 ORDER BY id LIMIT 1").fetchone()
    if not admin:
        db.execute(
            "INSERT INTO users (email, name, hashed_password, is_admin) VALUES (?, ?, ?, 1)",
            (ADMIN_EMAIL, "Admin", auth.hash_password(ADMIN_PASSWORD)),
        )
        db.commit()
        admin = db.execute("SELECT id FROM users WHERE is_admin = 1 ORDER BY id LIMIT 1").fetchone()
    # assign surveys that predate multi-user to the admin
    db.execute("UPDATE surveys SET owner_id = ? WHERE owner_id IS NULL", (admin["id"],))
    db.commit()


# --- auth helpers ---

def _owned_survey(db, slug: str, user):
    """The survey row if `user` may manage it (owner or admin), else None."""
    row = db.execute("SELECT * FROM surveys WHERE slug = ?", (slug,)).fetchone()
    if not row:
        return None
    if user["is_admin"] or row["owner_id"] == user["id"]:
        return row
    return None


# --- auth routes ---

@app.get("/login")
async def login_page(error: int = 0):
    # the login form lives on the combined landing page
    return RedirectResponse("/?error=1" if error else "/", status_code=302)


@app.post("/login")
async def login(email: str = Form(...), password: str = Form(...)):
    # In gateway mode the app switches its own login off rather than trusting
    # the proxy to hide it: two sets of credentials for one tool is exactly what
    # the SSO is there to remove.
    if auth.gateway_mode():
        return RedirectResponse("/", status_code=302)
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE email = ? AND is_active = 1", (email.strip().lower(),)
    ).fetchone()
    db.close()
    if not user or not auth.verify_password(password, user["hashed_password"]):
        return RedirectResponse("/?error=1", status_code=302)
    # password ok → pending session; full access only after the 2FA step
    response = RedirectResponse("/2fa", status_code=302)
    auth.set_session(response, user["id"], "pending_2fa")
    return response


@app.get("/register")
async def register_page():
    return RedirectResponse("/?tab=register", status_code=302)


@app.post("/register")
async def register(name: str = Form(...), email: str = Form(...), password: str = Form(...)):
    # Open registration is a `local` affordance. Behind the gate, who gets an
    # account is the gate's decision and arriving here would make a second one.
    if auth.gateway_mode():
        return RedirectResponse("/", status_code=302)
    email = email.strip().lower()
    name = name.strip()
    if not EMAIL_RE.match(email):
        return RedirectResponse("/?tab=register&reg_error=email", status_code=302)
    if len(password) < 8:
        return RedirectResponse("/?tab=register&reg_error=pwd", status_code=302)
    if not name:
        return RedirectResponse("/?tab=register&reg_error=name", status_code=302)
    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (email, name, hashed_password) VALUES (?, ?, ?)",
            (email, name, auth.hash_password(password)),
        )
        db.commit()
    except sqlite3.IntegrityError:
        db.close()
        return RedirectResponse("/?tab=register&reg_error=taken", status_code=302)
    user_id = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()["id"]
    db.close()
    # accounts are active immediately, but 2FA enrolment is mandatory before use
    response = RedirectResponse("/2fa", status_code=302)
    auth.set_session(response, user_id, "pending_2fa")
    return response


BORANT_LOGOUT_URL = os.getenv("BORANT_LOGOUT_URL", "https://id.borant.eu/logout")


@app.get("/logout")
async def logout():
    # In gateway mode dropping the local cookie is not signing out: the gate
    # still holds the session, and the next click walks straight back in.
    target = BORANT_LOGOUT_URL if auth.gateway_mode() else "/login"
    response = RedirectResponse(target, status_code=302)
    response.delete_cookie("session")
    return response


# --- two-factor (TOTP, mandatory) ---

@app.get("/2fa", response_class=HTMLResponse)
async def twofa_page(request: Request):
    db = get_db()
    if auth.current_user(request, db):   # already fully authenticated
        db.close()
        return RedirectResponse("/admin", status_code=302)
    user = auth.pending_user(request, db)
    db.close()
    if not user:                         # no pending session → back to login
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse(request, "twofa.html", {
        "enrolled": bool(user["totp_enabled"]),
        "email": user["email"],
    })


@app.post("/api/2fa/setup")
async def api_2fa_setup(request: Request):
    """Generate a secret + QR for enrolment (does not enable 2FA until confirmed)."""
    db = get_db()
    user = auth.pending_user(request, db)
    if not user:
        db.close()
        return JSONResponse({"error": "session expired"}, status_code=401)
    secret = totp.generate_secret()
    db.execute("UPDATE users SET totp_secret_encrypted = ? WHERE id = ?",
               (crypto.encrypt(secret), user["id"]))
    db.commit()
    db.close()
    uri = totp.provisioning_uri(secret, user["email"])
    return JSONResponse({"secret": secret, "uri": uri, "qr": totp.qr_data_uri(uri)})


@app.post("/api/2fa/confirm")
async def api_2fa_confirm(request: Request):
    db = get_db()
    user = auth.pending_user(request, db)
    if not user:
        db.close()
        return JSONResponse({"error": "session expired"}, status_code=401)
    if not user["totp_secret_encrypted"]:
        db.close()
        return JSONResponse({"error": "start the setup first"}, status_code=400)
    body = await request.json()
    if not totp.verify(crypto.decrypt(user["totp_secret_encrypted"]), body.get("code", "")):
        db.close()
        return JSONResponse({"error": "Invalid code — check your authenticator app"}, status_code=400)
    plain, hashes = totp.generate_backup_codes()
    db.execute("UPDATE users SET totp_enabled = 1, backup_codes_json = ? WHERE id = ?",
               (json.dumps(hashes), user["id"]))
    db.commit()
    db.close()
    response = JSONResponse({"ok": True, "backup_codes": plain})
    auth.set_session(response, user["id"], "full")
    return response


@app.post("/api/2fa/verify")
async def api_2fa_verify(request: Request):
    db = get_db()
    user = auth.pending_user(request, db)
    if not user:
        db.close()
        return JSONResponse({"error": "session expired"}, status_code=401)
    if not (user["totp_enabled"] and user["totp_secret_encrypted"]):
        db.close()
        return JSONResponse({"error": "2FA is not configured"}, status_code=400)
    body = await request.json()
    code = body.get("code", "")
    ok = totp.verify(crypto.decrypt(user["totp_secret_encrypted"]), code)
    if not ok:  # fall back to a one-time backup code
        remaining = totp.consume_backup_code(code, json.loads(user["backup_codes_json"] or "[]"))
        if remaining is not None:
            db.execute("UPDATE users SET backup_codes_json = ? WHERE id = ?",
                       (json.dumps(remaining), user["id"]))
            db.commit()
            ok = True
    db.close()
    if not ok:
        return JSONResponse({"error": "Invalid code"}, status_code=400)
    response = JSONResponse({"ok": True})
    auth.set_session(response, user["id"], "full")
    return response


# --- randomization ---

def _pool_counts(db, pool_id: int) -> tuple[dict, dict]:
    """(completed, pending) counts per condition for one pool.

    Balancing is driven by completed responses, with assignments issued in the
    last PENDING_ASSIGNMENT_MINUTES counted too. Without the pending term a burst
    of simultaneous starts would all see the same counts and land on the same
    arm; with it, an abandoned session stops distorting the arms once it ages out.
    """
    completed, pending = {}, {}
    rows = db.execute(
        f"""SELECT condition_key,
                   SUM(completed) AS done,
                   SUM(CASE WHEN completed = 0
                             AND issued_at > datetime('now', '-{PENDING_ASSIGNMENT_MINUTES} minutes')
                            THEN 1 ELSE 0 END) AS live
            FROM assignments WHERE pool_id = ? GROUP BY condition_key""",
        (pool_id,),
    ).fetchall()
    for r in rows:
        completed[r["condition_key"]] = r["done"] or 0
        pending[r["condition_key"]] = r["live"] or 0
    return completed, pending


def _assign_condition(db, pool_id: int, pool: list, show_count: int) -> tuple[list, int]:
    """Pick the least-used combination and record a pending assignment.

    Returns (pages, assignment_id). The assignment is marked completed only when
    the response is submitted, so page loads that go nowhere do not consume an arm.
    """
    all_conditions = [
        ",".join(sorted(combo))
        for combo in itertools.combinations(pool, show_count)
    ]
    completed, pending = _pool_counts(db, pool_id)
    load = {c: completed.get(c, 0) + pending.get(c, 0) for c in all_conditions}
    min_count = min(load.values(), default=0)
    chosen = random.choice([c for c in all_conditions if load[c] == min_count])
    cur = db.execute(
        "INSERT INTO assignments (pool_id, condition_key) VALUES (?, ?)",
        (pool_id, chosen),
    )
    assignment_id = cur.lastrowid
    # housekeeping: pending rows older than a week can never complete
    db.execute(
        "DELETE FROM assignments WHERE pool_id = ? AND completed = 0 "
        "AND issued_at < datetime('now', '-7 days')",
        (pool_id,),
    )
    db.commit()
    return chosen.split(","), assignment_id


def _complete_assignments(db, survey_id: int, ids) -> None:
    """Mark the pending assignments for a submitted response as completed.

    Ids come from the client, so the update is scoped to pools belonging to this
    survey: an arbitrary id from elsewhere cannot be flipped.
    """
    clean = [int(i) for i in ids if str(i).isdigit()][:10]
    if not clean:
        return
    placeholders = ",".join("?" * len(clean))
    db.execute(
        f"""UPDATE assignments SET completed = 1
            WHERE id IN ({placeholders})
              AND pool_id IN (SELECT id FROM rand_pools WHERE survey_id = ?)""",
        (*clean, survey_id),
    )


# --- panel recruitment helpers ---

def _panel_config(row) -> dict | None:
    """Panel settings for a survey row, or None when panel mode is off.
    Panel mode is on as soon as a token parameter is configured."""
    try:
        param = row["panel_token_param"]
    except (IndexError, KeyError):
        return None
    if not param:
        return None
    return {
        "param": param,
        "complete": row["panel_complete_url"] or "",
        "screenout": row["panel_screenout_url"] or "",
        "quotafull": row["panel_quotafull_url"] or "",
    }


def _read_panel_token(query_params, param: str) -> str | None:
    """The token from the entry URL, matched case-insensitively because
    providers disagree on capitalisation (RID, rid, Rid all appear in the wild).
    Anything outside the safe charset is treated as absent."""
    wanted = param.lower()
    for key, value in query_params.items():
        if key.lower() == wanted:
            value = (value or "").strip()
            return value if PANEL_TOKEN_RE.match(value) else None
    return None


def _panel_redirect(config: dict, outcome: str, token: str | None) -> str | None:
    """Return URL for an outcome, with the token substituted or appended.

    A `{token}` placeholder anywhere in the configured URL is replaced in place,
    which covers providers whose return URL carries the id mid-path. Otherwise
    the token is appended as a query parameter under the same name it arrived in.
    """
    url = (config.get(outcome) or "").strip()
    if not url:
        return None
    quoted = urllib.parse.quote(token or "", safe="")
    if "{token}" in url:
        return url.replace("{token}", quoted)
    if not token:
        return url
    sep = "&" if "?" in url else "?"
    return f"{url}{sep}{urllib.parse.quote(config['param'], safe='')}={quoted}"


# --- public routes ---

@app.get("/", response_class=HTMLResponse)
async def root(request: Request, error: int = 0, reg_error: str = "", tab: str = ""):
    db = get_db()
    user = auth.current_user(request, db)
    db.close()
    if user:
        return RedirectResponse("/admin", status_code=302)
    return templates.TemplateResponse(request, "landing.html", {
        "error": error,
        "reg_error": reg_error,
        "tab": "register" if (tab == "register" or reg_error) else "login",
    })


def _panel_stop(request, heading: str, message: str, redirect: str | None, status: int = 200):
    """Dead-end page for a panel respondent who cannot take the survey.
    Redirects to the provider when a return URL is configured, so the provider
    can classify them instead of recording an abandonment."""
    if redirect:
        return RedirectResponse(redirect, status_code=302)
    return templates.TemplateResponse(
        request,
        "panel_stop.html",
        {"heading": heading, "message": message},
        status_code=status,
    )


@app.get("/s/{slug}", response_class=HTMLResponse)
async def survey_page(request: Request, slug: str):
    db = get_db()
    row = db.execute(
        "SELECT * FROM surveys WHERE slug = ? AND active = 1", (slug,)
    ).fetchone()
    if not row:
        db.close()
        return templates.TemplateResponse(request, "closed.html", {}, status_code=404)

    # --- panel gate ---
    panel = _panel_config(row)
    panel_token = None
    if panel:
        panel_token = _read_panel_token(request.query_params, panel["param"])
        # owners can walk their own survey without a provider token
        preview = False
        if request.query_params.get("preview") == "1":
            viewer = auth.current_user(request, db)
            preview = bool(viewer and _owned_survey(db, slug, viewer))
        if not panel_token and not preview:
            db.close()
            return _panel_stop(
                request,
                "Invalid link",
                "This survey can only be entered through the link supplied by your panel "
                "provider, which carries the identifier needed to credit your participation. "
                "Please return to the panel and start again from there.",
                _panel_redirect(panel, "screenout", None),
                status=400,
            )
        if panel_token and db.execute(
            "SELECT 1 FROM responses WHERE survey_id = ? AND panel_token = ?",
            (row["id"], panel_token),
        ).fetchone():
            db.close()
            return _panel_stop(
                request,
                "Already completed",
                "Our records show this invitation has already been used to complete the "
                "survey. Each invitation can be used once.",
                _panel_redirect(panel, "screenout", panel_token),
            )

    assigned_pages_list: list = []
    pool_pages_list: list = []
    assignment_ids: list = []
    conditions: dict = {}
    page_orders: list = []
    pools = db.execute(
        "SELECT id, pool_pages, show_count, condition_var, condition_map, page_order "
        "FROM rand_pools WHERE survey_id = ? ORDER BY pool_order",
        (row["id"],),
    ).fetchall()
    for p in pools:
        pool = json.loads(p["pool_pages"])
        sc = p["show_count"]
        if pool and 0 < sc <= len(pool):
            pool_pages_list.extend(pool)
            assigned, assignment_id = _assign_condition(db, p["id"], pool, sc)
            assignment_ids.append(assignment_id)
            assigned_pages_list.extend(assigned)
            # condition variable: only meaningful when show_count=1
            cvar = p["condition_var"]
            if cvar and sc == 1 and len(assigned) == 1:
                page = assigned[0]
                cmap_raw = p["condition_map"]
                if cmap_raw:
                    try:
                        cmap = json.loads(cmap_raw)
                        value = cmap.get(page, page)
                    except (json.JSONDecodeError, TypeError):
                        value = page
                else:
                    value = page
                conditions[cvar] = value
                # optional page reordering, keyed by the condition value just
                # assigned. Only the sequence matters: the named pages are put
                # back into the slots they already occupy, in the order given,
                # so a pool can counterbalance presentation order without
                # duplicating pages (and question names) in the schema.
                porder_raw = p["page_order"]
                if porder_raw:
                    try:
                        pmap = json.loads(porder_raw)
                    except (json.JSONDecodeError, TypeError):
                        pmap = None
                    if isinstance(pmap, dict) and isinstance(pmap.get(value), list):
                        page_orders.append({"var": cvar, "pages": pmap[value]})
    assigned_pages = assigned_pages_list or None
    pool_pages = pool_pages_list or None
    db.close()

    return templates.TemplateResponse(request, "survey.html", {
        "title": row["title"],
        "slug": slug,
        "schema": json.loads(row["schema_json"]),
        "assigned_pages": assigned_pages,
        "pool_pages": pool_pages,
        "assignment_ids": assignment_ids,
        "conditions": conditions,
        "page_orders": page_orders,
        "panel_token": panel_token,
    })


@app.post("/s/{slug}/submit")
async def submit(slug: str, request: Request):
    db = get_db()
    row = db.execute("SELECT * FROM surveys WHERE slug = ?", (slug,)).fetchone()
    if not row or not row["active"]:
        db.close()
        return JSONResponse({"error": "survey not found or closed"}, status_code=404)
    data = await request.json()

    panel = _panel_config(row)
    token = None
    if panel:
        raw = str(data.get("_panel_token") or "").strip()
        token = raw if PANEL_TOKEN_RE.match(raw) else None
        data["_panel_token"] = token

    # outcome drives which return URL the respondent bounces to; the schema sets
    # it with a SurveyJS trigger (see the Panel section of the manage page)
    outcome = str(data.get("_outcome") or "complete").strip().lower()
    if outcome not in PANEL_OUTCOMES:
        outcome = "complete"

    try:
        db.execute(
            "INSERT INTO responses (survey_id, response_json, panel_token) VALUES (?, ?, ?)",
            (row["id"], json.dumps(data, ensure_ascii=False), token),
        )
    except sqlite3.IntegrityError:
        # the unique index caught a token being spent twice
        db.close()
        return JSONResponse(
            {"error": "duplicate", "redirect": _panel_redirect(panel, "screenout", token)},
            status_code=409,
        )
    _complete_assignments(db, row["id"], data.get("_assignment_ids") or [])
    db.commit()
    db.close()
    # No token means nobody to credit, so an owner previewing the questionnaire
    # is not bounced onto the provider's completion endpoint.
    return JSONResponse({
        "ok": True,
        "redirect": _panel_redirect(panel, outcome, token) if (panel and token) else None,
    })


# --- admin routes ---

@app.get("/admin", response_class=HTMLResponse)
async def admin_home(request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if user["is_admin"]:
        surveys = db.execute("""
            SELECT s.*, COUNT(r.id) AS response_count, u.email AS owner_email
            FROM surveys s
            LEFT JOIN responses r ON r.survey_id = s.id
            LEFT JOIN users u ON u.id = s.owner_id
            GROUP BY s.id
            ORDER BY s.created_at DESC
        """).fetchall()
    else:
        surveys = db.execute("""
            SELECT s.*, COUNT(r.id) AS response_count, NULL AS owner_email
            FROM surveys s
            LEFT JOIN responses r ON r.survey_id = s.id
            WHERE s.owner_id = ?
            GROUP BY s.id
            ORDER BY s.created_at DESC
        """, (user["id"],)).fetchall()
    db.close()
    keys = db.execute(
        "SELECT id, name, key, active, created_at, last_used_at FROM api_keys "
        "WHERE user_id = ? ORDER BY id DESC", (user["id"],)).fetchall()
    db.close()
    return templates.TemplateResponse(request, "admin.html", {
        "surveys": surveys, "user": user, "api_keys": keys,
        "public_url": os.environ.get("PUBLIC_URL", "").rstrip("/"),
    })


# --- MCP keys ---
# Minted here rather than handed out by an administrator, because a key is a
# credential of the person holding it: it reaches exactly their surveys, so
# there is nothing to approve.

@app.post("/admin/keys")
async def mint_key(request: Request, name: str = Form("")):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    db.execute("INSERT INTO api_keys (user_id, name, key) VALUES (?, ?, ?)",
               (user["id"], (name or "").strip()[:60] or "MCP key", auth.new_api_key()))
    db.commit()
    db.close()
    return RedirectResponse("/admin", status_code=302)


@app.post("/admin/keys/{key_id}/revoke")
async def revoke_key(key_id: int, request: Request):
    """Revoked, never deleted: the row is what tells you a key existed and when
    it was last used, which is the only trace of where it might still be
    configured."""
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    db.execute("UPDATE api_keys SET active = 0 WHERE id = ? AND user_id = ?",
               (key_id, user["id"]))
    db.commit()
    db.close()
    return RedirectResponse("/admin", status_code=302)


@app.post("/admin/surveys")
async def create_survey(
    request: Request,
    title: str = Form(...),
    slug: str = Form(...),
    schema_file: UploadFile = File(None),
    schema_text: str = Form(""),
):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)

    if schema_file and schema_file.filename:
        raw = await schema_file.read()
        schema_str = raw.decode("utf-8")
    elif schema_text.strip():
        schema_str = schema_text.strip()
    else:
        db.close()
        return RedirectResponse("/admin?error=no_schema", status_code=302)

    try:
        json.loads(schema_str)
    except json.JSONDecodeError:
        db.close()
        return RedirectResponse("/admin?error=invalid_json", status_code=302)

    slug = slug.strip().lower().replace(" ", "-")

    try:
        db.execute(
            "INSERT INTO surveys (slug, title, schema_json, owner_id) VALUES (?, ?, ?, ?)",
            (slug, title, schema_str, user["id"]),
        )
        db.commit()
    except sqlite3.IntegrityError:
        db.close()
        return RedirectResponse("/admin?error=duplicate_slug", status_code=302)
    db.close()
    return RedirectResponse("/admin", status_code=302)


@app.get("/admin/surveys/{slug}", response_class=HTMLResponse)
async def manage_survey(slug: str, request: Request):
    """Per-survey hub: share link + QR, exports, questionnaire tools,
    configuration links, and the danger zone."""
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    row = _owned_survey(db, slug, user)
    if not row:
        db.close()
        return RedirectResponse("/admin", status_code=302)
    stats = db.execute(
        "SELECT COUNT(*) AS n, MAX(submitted_at) AS last FROM responses WHERE survey_id = ?",
        (row["id"],),
    ).fetchone()
    pool_count = db.execute(
        "SELECT COUNT(*) FROM rand_pools WHERE survey_id = ?", (row["id"],)
    ).fetchone()[0]
    owner = db.execute(
        "SELECT email FROM users WHERE id = ?", (row["owner_id"],)
    ).fetchone()
    db.close()

    schema = json.loads(row["schema_json"])
    locales = review_export._locales_in(schema) or {"en"}
    langs = [l for l in review_export.LANGS if l in locales]

    upload_dir = os.path.join(UPLOADS_PATH, slug)
    files_count = len(os.listdir(upload_dir)) if os.path.isdir(upload_dir) else 0

    # honour the reverse proxy's scheme so the QR points at the public URL
    scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
    host = request.headers.get("host", request.url.netloc)
    public_url = f"{scheme}://{host}/s/{slug}"

    panel = _panel_config(row) or {}
    return templates.TemplateResponse(request, "manage.html", {
        "user": user,
        "slug": slug,
        "title": row["title"],
        "active": row["active"],
        "created_at": row["created_at"],
        "owner_email": owner["email"] if owner else None,
        "response_count": stats["n"],
        "last_response": stats["last"],
        "pool_count": pool_count,
        "files_count": files_count,
        "langs": langs,
        "public_url": public_url,
        "qr": totp.qr_data_uri(public_url),
        "panel": panel,
        "panel_error": request.query_params.get("panel_error", ""),
        "panel_saved": request.query_params.get("panel_saved") == "1",
    })


@app.post("/admin/surveys/{slug}/panel")
async def save_panel_config(
    slug: str,
    request: Request,
    token_param: str = Form(""),
    complete_url: str = Form(""),
    screenout_url: str = Form(""),
    quotafull_url: str = Form(""),
):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    row = _owned_survey(db, slug, user)
    if not row:
        db.close()
        return RedirectResponse("/admin", status_code=302)

    token_param = token_param.strip()
    urls = {k: v.strip() for k, v in (
        ("complete", complete_url), ("screenout", screenout_url), ("quotafull", quotafull_url)
    )}

    def fail(message: str):
        db.close()
        return RedirectResponse(
            f"/admin/surveys/{slug}?panel_error={urllib.parse.quote(message)}",
            status_code=302,
        )

    if token_param and not PANEL_PARAM_RE.match(token_param):
        return fail("The token parameter must be 1 to 32 letters, digits, hyphens or underscores.")
    for name, url in urls.items():
        if url and not url.lower().startswith(("http://", "https://")):
            return fail(f"The {name} URL must start with http:// or https://.")
    if not token_param and any(urls.values()):
        return fail("Set a token parameter to switch panel mode on, or clear the return URLs.")

    db.execute(
        """UPDATE surveys SET panel_token_param = ?, panel_complete_url = ?,
                              panel_screenout_url = ?, panel_quotafull_url = ?
           WHERE id = ?""",
        (token_param or None, urls["complete"] or None,
         urls["screenout"] or None, urls["quotafull"] or None, row["id"]),
    )
    db.commit()
    db.close()
    return RedirectResponse(f"/admin/surveys/{slug}?panel_saved=1", status_code=302)


@app.get("/admin/surveys/{slug}/edit", response_class=HTMLResponse)
async def edit_survey_page(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    row = _owned_survey(db, slug, user)
    db.close()
    if not row:
        return RedirectResponse("/admin", status_code=302)
    return templates.TemplateResponse(request, "edit.html", {
        "slug": slug,
        "title": row["title"],
        "schema_json": json.dumps(json.loads(row["schema_json"]), indent=2, ensure_ascii=False),
    })


@app.post("/admin/surveys/{slug}/edit")
async def edit_survey(
    slug: str,
    request: Request,
    title: str = Form(...),
    schema_file: UploadFile = File(None),
    schema_text: str = Form(""),
):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if not _owned_survey(db, slug, user):
        db.close()
        return RedirectResponse("/admin", status_code=302)

    if schema_file and schema_file.filename:
        raw = await schema_file.read()
        schema_str = raw.decode("utf-8")
    elif schema_text.strip():
        schema_str = schema_text.strip()
    else:
        db.close()
        return RedirectResponse(f"/admin/surveys/{slug}/edit?error=no_schema", status_code=302)

    try:
        json.loads(schema_str)
    except json.JSONDecodeError:
        db.close()
        return RedirectResponse(f"/admin/surveys/{slug}/edit?error=invalid_json", status_code=302)

    db.execute(
        "UPDATE surveys SET title = ?, schema_json = ? WHERE slug = ?",
        (title, schema_str, slug),
    )
    db.commit()
    db.close()
    return RedirectResponse(f"/admin/surveys/{slug}", status_code=302)


@app.post("/admin/surveys/{slug}/toggle")
async def toggle_survey(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if _owned_survey(db, slug, user):
        db.execute("UPDATE surveys SET active = 1 - active WHERE slug = ?", (slug,))
        db.commit()
    db.close()
    return RedirectResponse("/admin", status_code=302)


@app.post("/admin/surveys/{slug}/delete")
async def delete_survey(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if _owned_survey(db, slug, user):
        db.execute("DELETE FROM surveys WHERE slug = ?", (slug,))
        db.commit()
        shutil.rmtree(os.path.join(UPLOADS_PATH, slug), ignore_errors=True)
    db.close()
    return RedirectResponse("/admin", status_code=302)


# --- admin randomization ---

@app.get("/admin/surveys/{slug}/randomization", response_class=HTMLResponse)
async def randomization_page(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    survey = _owned_survey(db, slug, user)
    if not survey:
        db.close()
        return RedirectResponse("/admin", status_code=302)
    schema = json.loads(survey["schema_json"])
    page_names = [p.get("name", f"page{i+1}") for i, p in enumerate(schema.get("pages", []))]
    pools_raw = db.execute(
        "SELECT id, pool_name, pool_order, pool_pages, show_count, condition_var, "
        "condition_map, page_order FROM rand_pools WHERE survey_id = ? ORDER BY pool_order",
        (survey["id"],),
    ).fetchall()
    pools_data = []
    for pool in pools_raw:
        completed, pending = _pool_counts(db, pool["id"])
        keys = sorted(set(completed) | set(pending))
        counts = [{
            "condition_key": k,
            "count": completed.get(k, 0),
            "pending": pending.get(k, 0),
        } for k in keys]
        pools_data.append({
            "id": pool["id"],
            "pool_name": pool["pool_name"],
            "pool_pages": json.loads(pool["pool_pages"]),
            "show_count": pool["show_count"],
            "condition_var": pool["condition_var"] or "",
            "condition_map": pool["condition_map"] or "",
            "page_order": pool["page_order"] or "",
            "counts": counts,
            "total": sum(c["count"] for c in counts),
            "pending_total": sum(c["pending"] for c in counts),
        })
    db.close()
    return templates.TemplateResponse(request, "randomization.html", {
        "slug": slug,
        "title": survey["title"],
        "page_names": page_names,
        "pools": pools_data,
    })


@app.post("/admin/surveys/{slug}/randomization/add-pool")
async def add_pool(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    row = _owned_survey(db, slug, user)
    if row:
        order = db.execute(
            "SELECT COALESCE(MAX(pool_order)+1, 0) FROM rand_pools WHERE survey_id = ?", (row["id"],)
        ).fetchone()[0]
        db.execute(
            "INSERT INTO rand_pools (survey_id, pool_name, pool_order) VALUES (?, ?, ?)",
            (row["id"], f"Pool {order + 1}", order),
        )
        db.commit()
    db.close()
    return RedirectResponse(f"/admin/surveys/{slug}/randomization", status_code=302)


def _pool_belongs(db, slug: str, pool_id: int, user) -> bool:
    survey = _owned_survey(db, slug, user)
    if not survey:
        return False
    row = db.execute("SELECT survey_id FROM rand_pools WHERE id = ?", (pool_id,)).fetchone()
    return bool(row and row["survey_id"] == survey["id"])


@app.post("/admin/surveys/{slug}/randomization/{pool_id}/save")
async def save_pool(slug: str, pool_id: int, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if not _pool_belongs(db, slug, pool_id, user):
        db.close()
        return RedirectResponse("/admin", status_code=302)
    form = await request.form()
    pool_pages = form.getlist("pool_pages")
    pool_name = form.get("pool_name", "Pool").strip() or "Pool"
    try:
        show_count = max(1, int(form.get("show_count", 1)))
    except ValueError:
        show_count = 1
    condition_var = form.get("condition_var", "").strip() or None
    # condition_map and page_order only valid when show_count=1; validate JSON
    # if provided
    condition_map = None
    page_order = None
    if show_count == 1 and condition_var:
        cmap_raw = form.get("condition_map", "").strip()
        if cmap_raw:
            try:
                json.loads(cmap_raw)
                condition_map = cmap_raw
            except json.JSONDecodeError:
                pass  # silently discard invalid JSON
        porder_raw = form.get("page_order", "").strip()
        if porder_raw:
            try:
                parsed = json.loads(porder_raw)
                # {condition value: [page names in the order they should appear]}
                if isinstance(parsed, dict) and all(
                    isinstance(v, list) for v in parsed.values()
                ):
                    page_order = porder_raw
            except json.JSONDecodeError:
                pass  # silently discard invalid JSON
    db.execute(
        "UPDATE rand_pools SET pool_name = ?, pool_pages = ?, show_count = ?, "
        "condition_var = ?, condition_map = ?, page_order = ? WHERE id = ?",
        (pool_name, json.dumps(pool_pages), show_count, condition_var,
         condition_map, page_order, pool_id),
    )
    # a configuration change invalidates the balance history for this pool
    db.execute("DELETE FROM assignments WHERE pool_id = ?", (pool_id,))
    db.execute("DELETE FROM assignment_counts WHERE pool_id = ?", (pool_id,))
    db.commit()
    db.close()
    return RedirectResponse(f"/admin/surveys/{slug}/randomization", status_code=302)


@app.post("/admin/surveys/{slug}/randomization/{pool_id}/delete")
async def delete_pool(slug: str, pool_id: int, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if _pool_belongs(db, slug, pool_id, user):
        db.execute("DELETE FROM rand_pools WHERE id = ?", (pool_id,))
        db.commit()
    db.close()
    return RedirectResponse(f"/admin/surveys/{slug}/randomization", status_code=302)


@app.post("/admin/surveys/{slug}/randomization/{pool_id}/reset")
async def reset_pool_counts(slug: str, pool_id: int, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if _pool_belongs(db, slug, pool_id, user):
        db.execute("DELETE FROM assignments WHERE pool_id = ?", (pool_id,))
        db.execute("DELETE FROM assignment_counts WHERE pool_id = ?", (pool_id,))
        db.commit()
    db.close()
    return RedirectResponse(f"/admin/surveys/{slug}/randomization", status_code=302)


# --- file uploads ---

def _safe_filename(name: str) -> str:
    name = os.path.basename(name)
    name = re.sub(r"[^\w.\-]", "_", name)
    return name or "file"


def _upload_dir(slug: str) -> str:
    path = os.path.join(UPLOADS_PATH, slug)
    os.makedirs(path, exist_ok=True)
    return path


@app.get("/admin/surveys/{slug}/files", response_class=HTMLResponse)
async def files_page(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    row = _owned_survey(db, slug, user)
    db.close()
    if not row:
        return RedirectResponse("/admin", status_code=302)
    d = _upload_dir(slug)
    files = sorted(os.listdir(d))
    return templates.TemplateResponse(request, "files.html", {
        "slug": slug,
        "title": row["title"],
        "files": files,
    })


@app.post("/admin/surveys/{slug}/upload")
async def upload_file(slug: str, request: Request, file: UploadFile = File(...)):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    owned = _owned_survey(db, slug, user)
    db.close()
    if not owned:
        return RedirectResponse("/admin", status_code=302)

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return RedirectResponse(f"/admin/surveys/{slug}/files?error=ext", status_code=302)

    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        return RedirectResponse(f"/admin/surveys/{slug}/files?error=size", status_code=302)

    dest = os.path.join(_upload_dir(slug), _safe_filename(file.filename))
    with open(dest, "wb") as f:
        f.write(content)

    return RedirectResponse(f"/admin/surveys/{slug}/files", status_code=302)


@app.post("/admin/surveys/{slug}/files/{filename}/delete")
async def delete_file(slug: str, filename: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    owned = _owned_survey(db, slug, user)
    db.close()
    if not owned:
        return RedirectResponse("/admin", status_code=302)
    safe = _safe_filename(filename)
    path = os.path.join(UPLOADS_PATH, slug, safe)
    if os.path.isfile(path):
        os.remove(path)
    return RedirectResponse(f"/admin/surveys/{slug}/files", status_code=302)


# --- export ---

def _flatten(data: dict, prefix: str = "") -> dict:
    result = {}
    for k, v in data.items():
        key = f"{prefix}_{k}" if prefix else k
        if isinstance(v, dict):
            result.update(_flatten(v, key))
        elif isinstance(v, list):
            result[key] = ";".join(str(i) for i in v)
        else:
            result[key] = v
    return result


def _get_responses(db, slug: str):
    row = db.execute("SELECT id FROM surveys WHERE slug = ?", (slug,)).fetchone()
    if not row:
        return None
    return db.execute(
        "SELECT response_json, submitted_at FROM responses WHERE survey_id = ? ORDER BY submitted_at",
        (row["id"],),
    ).fetchall()


def _flat_table(rows):
    """(columns, flat_rows) shared by the CSV and Excel exports, so the two
    formats always agree on structure. Column order follows first appearance;
    _submitted_at goes last."""
    flat_rows = []
    all_keys: list[str] = []
    seen_keys: set[str] = set()
    for r in rows:
        flat = _flatten(json.loads(r["response_json"]))
        flat["_submitted_at"] = r["submitted_at"]
        flat_rows.append(flat)
        for k in flat:
            if k not in seen_keys:
                seen_keys.add(k)
                all_keys.append(k)
    cols = [k for k in all_keys if k != "_submitted_at"] + ["_submitted_at"]
    return cols, flat_rows


def _build_xlsx(cols, flat_rows) -> bytes:
    """Excel workbook: frozen bold header, autofilter, native numeric types.
    Strings that look like formulas are forced to text — open-ended answers
    must never execute in a reviewer's Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"
    for j, k in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=k)
        cell.font = Font(bold=True)
    for i, flat in enumerate(flat_rows, start=2):
        for j, k in enumerate(cols, start=1):
            v = flat.get(k, "")
            cell = ws.cell(row=i, column=j)
            cell.value = v
            if isinstance(v, str) and v.startswith("="):
                cell.data_type = "s"
    for j, k in enumerate(cols, start=1):
        sample = [len(str(fr.get(k, ""))) for fr in flat_rows[:200]]
        ws.column_dimensions[get_column_letter(j)].width = min(60, max(10, len(k), *sample))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/admin/surveys/{slug}/export.csv")
async def export_csv(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if not _owned_survey(db, slug, user):
        db.close()
        return JSONResponse({"error": "not found"}, status_code=404)
    rows = _get_responses(db, slug)
    db.close()
    if rows is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    if not rows:
        return HTMLResponse("No responses yet.")

    cols, flat_rows = _flat_table(rows)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore", restval="")
    writer.writeheader()
    writer.writerows(flat_rows)

    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{slug}.csv"'},
    )


@app.get("/admin/surveys/{slug}/export.xlsx")
async def export_xlsx(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if not _owned_survey(db, slug, user):
        db.close()
        return JSONResponse({"error": "not found"}, status_code=404)
    rows = _get_responses(db, slug)
    db.close()
    if rows is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    if not rows:
        return HTMLResponse("No responses yet.")

    cols, flat_rows = _flat_table(rows)
    data = _build_xlsx(cols, flat_rows)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{slug}.xlsx"'},
    )


@app.get("/admin/surveys/{slug}/review.docx")
async def export_review_docx(slug: str, request: Request):
    """Word rendering of the questionnaire itself (not the responses), for
    circulating to reviewers: primary-language texts, answer formats,
    visibility logic, randomization pools, and translation coverage flags."""
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    row = _owned_survey(db, slug, user)
    if not row:
        db.close()
        return JSONResponse({"error": "not found"}, status_code=404)
    pool_rows = db.execute(
        "SELECT pool_name, pool_pages, show_count, condition_var, condition_map, page_order "
        "FROM rand_pools WHERE survey_id = ? ORDER BY pool_order",
        (row["id"],),
    ).fetchall()
    db.close()

    data = review_export.build_review_docx(
        json.loads(row["schema_json"]),
        review_export.pools_from_rows(pool_rows),
        row["title"],
    )
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{slug}-review.docx"'},
    )


@app.get("/admin/surveys/{slug}/export.json")
async def export_json(slug: str, request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if not _owned_survey(db, slug, user):
        db.close()
        return JSONResponse({"error": "not found"}, status_code=404)
    rows = _get_responses(db, slug)
    db.close()
    if rows is None:
        return JSONResponse({"error": "not found"}, status_code=404)

    data = [
        {"submitted_at": r["submitted_at"], "data": json.loads(r["response_json"])}
        for r in rows
    ]
    return StreamingResponse(
        io.BytesIO(json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{slug}.json"'},
    )


# --- profile (self-service) ---

@app.get("/profile", response_class=HTMLResponse)
async def profile_page(request: Request, ok: str = "", error: str = ""):
    db = get_db()
    user = auth.current_user(request, db)
    db.close()
    if not user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse(request, "profile.html", {
        "user": user, "ok": ok, "error": error,
    })


@app.post("/profile/password")
async def profile_password(
    request: Request,
    current: str = Form(...),
    new_password: str = Form(...),
):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if not auth.verify_password(current, user["hashed_password"]):
        db.close()
        return RedirectResponse("/profile?error=current", status_code=302)
    if len(new_password) < 8:
        db.close()
        return RedirectResponse("/profile?error=short", status_code=302)
    db.execute("UPDATE users SET hashed_password = ? WHERE id = ?",
               (auth.hash_password(new_password), user["id"]))
    db.commit()
    db.close()
    return RedirectResponse("/profile?ok=password", status_code=302)


@app.post("/api/profile/backup-codes")
async def regenerate_backup_codes(request: Request):
    db = get_db()
    user = auth.current_user(request, db)
    if not user:
        db.close()
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    if not user["totp_enabled"]:
        db.close()
        return JSONResponse({"error": "2FA not enabled"}, status_code=400)
    plain, hashes = totp.generate_backup_codes()
    db.execute("UPDATE users SET backup_codes_json = ? WHERE id = ?",
               (json.dumps(hashes), user["id"]))
    db.commit()
    db.close()
    return JSONResponse({"backup_codes": plain})


# --- admin: user management ---

@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request, tmp_uid: int = 0, tmp_password: str = ""):
    db = get_db()
    user = auth.current_user(request, db)
    if not user or not user["is_admin"]:
        db.close()
        return RedirectResponse("/admin" if user else "/login", status_code=302)
    users = db.execute("""
        SELECT u.*, COUNT(s.id) AS survey_count
        FROM users u
        LEFT JOIN surveys s ON s.owner_id = u.id
        GROUP BY u.id
        ORDER BY u.created_at
    """).fetchall()
    db.close()
    return templates.TemplateResponse(request, "admin_users.html", {
        "user": user, "users": users,
        "tmp_uid": tmp_uid, "tmp_password": tmp_password,
    })


def _admin_or_none(request, db):
    user = auth.current_user(request, db)
    if not user or not user["is_admin"]:
        return None
    return user


@app.post("/admin/users/{uid}/toggle-active")
async def admin_toggle_active(uid: int, request: Request):
    db = get_db()
    admin = _admin_or_none(request, db)
    if not admin:
        db.close()
        return RedirectResponse("/login", status_code=302)
    if uid != admin["id"]:  # never lock yourself out
        db.execute("UPDATE users SET is_active = 1 - is_active WHERE id = ?", (uid,))
        db.commit()
    db.close()
    return RedirectResponse("/admin/users", status_code=302)


@app.post("/admin/users/{uid}/reset-2fa")
async def admin_reset_2fa(uid: int, request: Request):
    """Clear a user's 2FA so they re-enrol on next login (lost-device recovery)."""
    db = get_db()
    admin = _admin_or_none(request, db)
    if not admin:
        db.close()
        return RedirectResponse("/login", status_code=302)
    db.execute(
        "UPDATE users SET totp_enabled = 0, totp_secret_encrypted = NULL, backup_codes_json = NULL WHERE id = ?",
        (uid,),
    )
    db.commit()
    db.close()
    return RedirectResponse("/admin/users", status_code=302)


@app.post("/admin/users/{uid}/reset-password")
async def admin_reset_password(uid: int, request: Request):
    """Set a fresh temporary password, shown once to the admin to hand over."""
    db = get_db()
    admin = _admin_or_none(request, db)
    if not admin:
        db.close()
        return RedirectResponse("/login", status_code=302)
    target = db.execute("SELECT id FROM users WHERE id = ?", (uid,)).fetchone()
    if not target:
        db.close()
        return RedirectResponse("/admin/users", status_code=302)
    temp = secrets.token_urlsafe(9)
    db.execute("UPDATE users SET hashed_password = ? WHERE id = ?",
               (auth.hash_password(temp), uid))
    db.commit()
    db.close()
    return RedirectResponse(f"/admin/users?tmp_uid={uid}&tmp_password={temp}", status_code=302)
